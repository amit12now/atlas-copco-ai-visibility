#!/usr/bin/env python3
"""
create_new_template.py  (Atlas Copco build)
--------------------------------------------
One-time (or "start fresh") utility that builds a new
AI_Visibility_Tracking_Template.xlsx from the raw exports sitting in a folder.

It looks for these exact filenames in --source-dir (defaults to the current
folder) and loads whatever it finds (all are optional):

  AI-overviews-SEMRUSH.xls                   -> RAW_AIO_Rankings
  brand_topics_atlascopco_com_us.csv         -> RAW_Brand_Topics
  prompts_atlascopco_com_us.csv              -> RAW_Prompts
  chatgpt_results_uae.csv                    -> RAW_ChatGPT_Results
  AI-Traffic-analytics.xlsx  (Sheet1)        -> RAW_GA4_Traffic
  gsc-AIO-Comparison.xlsx    (Pages)         -> RAW_GSC_Pages
  gsc-AIO-Comparison.xlsx    (Countries)     -> RAW_GSC_Countries
  gsc-AIO-Comparison.xlsx    (Devices)       -> RAW_GSC_Devices
  ("AIO" = AI Overview: this is Search Console's AI Overviews / AI Mode
  reporting, not the classic web-search Performance report. Falls back to
  AI-Performanc-GSC.xlsx if gsc-AIO-Comparison.xlsx isn't found.)

Note: there is no RAW_Gap_Topics / RAW_Sources source for this client (the
Gap Topics export we received is prompt-level, not topic-level, and no
Sources export was provided) -- those sheets are created empty on purpose.
The dashboard already tolerates empty Sources/Gap data gracefully.

It also creates the Settings, README, and Mapping_Reference sheets, and empty
(header-only) Clean_* sheets. After running this, run update_template.py to
populate the Clean_* sheets from the RAW_* data.

Usage:
  python create_new_template.py [--source-dir DIR] [--out FILE.xlsx]
                                 [--brand-domain DOMAIN] [--report-month YYYY-MM]
"""
import argparse
import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
RAW_HEADER_FILL = PatternFill("solid", start_color="843C0C", end_color="843C0C")
TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1F3864")
BODY_FONT = Font(name="Arial", size=10)
NOTE_FONT = Font(name="Arial", italic=True, size=9, color="808080")

CLEAN_SCHEMAS = {
    "Clean_AIO_Rankings": ["Report_Month", "Keyword", "Domain", "Is_Brand", "Position",
                            "Visibility_Pct", "Result_Type", "Landing_URL", "Position_Change",
                            "Visibility_Change", "Search_Volume", "Keyword_Difficulty", "CPC",
                            "Intent_Code", "Tags"],
    "Clean_Brand_Topics": ["Report_Month", "Topic", "Country", "Visibility", "Difficulty",
                            "Mentions", "Search_Volume", "Volume_Trend", "Intent_Task",
                            "Intent_Informational", "Intent_Navigational", "Intent_Commercial",
                            "Intent_Transactional"],
    "Clean_Gap_Topics": ["Report_Month", "Topic", "Country", "Visibility", "Difficulty",
                          "Brand_Mentions", "Search_Volume", "Volume_Trend", "Intent_Task",
                          "Intent_Informational", "Intent_Navigational", "Intent_Commercial",
                          "Intent_Transactional"],
    "Clean_Gap_Topics_Competitors": ["Report_Month", "Topic", "Competitor_Domain", "Competitor_Mentions"],
    "Clean_Prompts": ["Report_Month", "Prompt", "Country", "LLM", "Category", "Topic_ID",
                       "Topic_Name", "Topic_Volume", "Topic_Visibility", "Topic_Difficulty",
                       "Mentioned_Brands_Count", "Sources_Count", "Brief_Response",
                       "Topic_Intent_Task", "Topic_Intent_Informational", "Topic_Intent_Navigational",
                       "Topic_Intent_Commercial", "Topic_Intent_Transactional"],
    "Clean_Sources": ["Report_Month", "URL", "Country", "Prompts_Count"],
    "Clean_ChatGPT_Results": ["Report_Month", "Keyword", "Brand_Mentioned", "Source_URL_Count",
                               "Answer", "Source_URLs"],
    "Clean_GA4_Traffic": ["Year_Month", "Month_Label", "Is_To_Date", "Platform",
                           "Platform_Domain", "Sessions"],
    "Clean_GSC_Performance": ["Report_Month", "Dimension_Type", "Dimension_Value",
                               "Impressions_Last_28d", "Impressions_Prev_28d", "Change_Pct"],
}

MAPPING_REFERENCE = [
    ("RAW_AIO_Rankings", "Semrush > Position Tracking, AI Overview filter, exported as XLSX. Daily per-domain position/type/landing columns for the tracked period, plus one period-long 'difference' column per domain.", "Clean_AIO_Rankings", "Per-competitor, per-day column blocks are unpivoted into one row per Keyword x Domain x Day, then averaged across the period into one row per Keyword x Domain for Report_Month. '-' positions become blank. Domain inferred from the '*.domain/*_YYYYMMDD' column pattern, so it adapts automatically if competitors are added or removed."),
    ("RAW_Brand_Topics", "Semrush AI Visibility / Brand Monitoring > topics where the brand is mentioned, exported as CSV.", "Clean_Brand_Topics", "Direct column rename. The packed 'intents' string (task:1;informational:46;...) is split into 5 numeric Intent_* columns."),
    ("RAW_Gap_Topics", "(Not used for this client -- the gap export received was prompt-level, not topic-level. Sheet stays empty.)", "Clean_Gap_Topics + Clean_Gap_Topics_Competitors", "No transform run; dashboard handles empty gracefully (these sheets aren't rendered in the UI anyway)."),
    ("RAW_Prompts", "Semrush AI Visibility > Prompts detail (per-LLM prompt tracking), exported as CSV.", "Clean_Prompts", "Direct column rename; 'topic_intents' split into 5 numeric Topic_Intent_* columns."),
    ("RAW_Sources", "(Not used for this client -- no Sources export was provided. Sheet stays empty.)", "Clean_Sources", "No transform run; the Sources tab/section is hidden in the dashboard when this is empty."),
    ("RAW_ChatGPT_Results", "Direct ChatGPT query checker export (keyword, answer, source_urls), CSV.", "Clean_ChatGPT_Results", "Adds Brand_Mentioned (true if Settings!Brand_Domain appears in the answer or source URLs) and Source_URL_Count (urls are '|'-separated)."),
    ("RAW_GA4_Traffic", "GA4 > Acquisition report filtered/segmented to AI referral sources (ChatGPT, Gemini, Copilot), exported as XLSX, Month as rows and each platform as a column.", "Clean_GA4_Traffic", "Wide table unpivoted to one row per Month x Platform. A redundant 'Total AI (Top 3)' column is dropped. The current month's '...*' marker is parsed into an Is_To_Date flag."),
    ("RAW_GSC_Pages / Countries / Devices", "Google Search Console > Performance report (gsc-AIO-Comparison.xlsx), Pages/Countries/Devices tables, filtered to the /compressors site section and exported with two date-range Impressions columns (current period vs. prior period).", "Clean_GSC_Performance", "The three tables are stacked into one long table with a Dimension_Type column (Page/Country/Device). The two 'M/D/YY - M/D/YY Impressions' columns are matched by parsing their start dates -- the later range becomes Impressions_Last_28d, the earlier becomes Impressions_Prev_28d -- and Change_Pct is computed from them. Falls back to a single Impressions column (Change_Pct left blank) if an export without a comparison is used."),
]

README_LINES = [
    ("AI Visibility & SEO Tracking Template -- Atlas Copco", TITLE_FONT, None),
    ("", None, None),
    ("PURPOSE", HEADER_FONT, None),
    ("Tracks AI Overview rankings, AI-visibility (brand topics & prompts),", BODY_FONT, None),
    ("ChatGPT citation checks, GA4 AI-referral traffic, and Google Search Console", BODY_FONT, None),
    ("performance in one standardized workbook that feeds the Streamlit dashboard.", BODY_FONT, None),
    ("", None, None),
    ("SHEET COLOR KEY", HEADER_FONT, None),
    ("Orange tabs (RAW_*)  = paste raw exports here every month, overwriting old rows.", BODY_FONT, None),
    ("Blue tabs (Clean_*)  = built by the script. Do not edit by hand. The dashboard", BODY_FONT, None),
    ("                       reads ONLY these sheets, and they accumulate history", BODY_FONT, None),
    ("                       across months (one batch per Report_Month).", BODY_FONT, None),
    ("", None, None),
    ("MONTHLY UPDATE STEPS", HEADER_FONT, None),
    ("1. Export fresh data from Semrush (AI Overview tracking, AI Visibility brand", BODY_FONT, None),
    ("   topics, prompts), the ChatGPT checker, GA4, and Search Console.", BODY_FONT, None),
    ("2. For each RAW_* tab: select the old data rows (below the header) and delete", BODY_FONT, None),
    ("   them, then paste the new export so it starts directly under the existing", BODY_FONT, None),
    ("   header row.", BODY_FONT, None),
    ("3. On the Settings tab, set Report_Month to the new period, e.g. '2026-07'.", BODY_FONT, None),
    ("4. Save this file.", BODY_FONT, None),
    ("5. Run:  python update_template.py \"AI_Visibility_Tracking_Template.xlsx\"", BODY_FONT, None),
    ("6. Refresh the dashboard:  streamlit run dashboard.py", BODY_FONT, None),
    ("", None, None),
    ("NOTE: RAW_Sources and RAW_Gap_Topics are intentionally empty for this client", NOTE_FONT, None),
    ("(no Sources export was supplied, and the Gap Topics export received was", NOTE_FONT, None),
    ("prompt-level rather than topic-level). The dashboard hides the Sources tab", NOTE_FONT, None),
    ("when that data is empty.", NOTE_FONT, None),
    ("", None, None),
    ("See the Mapping_Reference tab for exactly how each RAW column becomes a Clean", BODY_FONT, None),
    ("column. See update_template.py for the executable version of that logic.", BODY_FONT, None),
]


def style_header_row(ws, fill):
    for cell in ws[1]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autosize(ws, ncols, width=18):
    for i in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(i)].width = width


def add_raw_sheet_from_df(wb, sheet_name, df):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "843C0C"
    if df is None or df.empty:
        return ws
    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=1, column=j, value=str(col))
    for i, row in enumerate(df.itertuples(index=False), start=2):
        for j, val in enumerate(row, start=1):
            v = None if pd.isna(val) else val
            ws.cell(row=i, column=j, value=v)
    style_header_row(ws, RAW_HEADER_FILL)
    autosize(ws, len(df.columns))
    return ws


def add_empty_raw_sheet(wb, sheet_name, headers):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "843C0C"
    for j, col in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=col)
    style_header_row(ws, RAW_HEADER_FILL)
    autosize(ws, len(headers))
    return ws


def add_clean_sheet(wb, sheet_name):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_properties.tabColor = "1F3864"
    headers = CLEAN_SCHEMAS[sheet_name]
    for j, col in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=col)
    style_header_row(ws, HEADER_FILL)
    autosize(ws, len(headers))
    return ws


def load_csv(path):
    if path.exists():
        return pd.read_csv(path)
    return None


def load_aio_xlsx(path):
    """Semrush AIO export: an OOXML/xlsx file (sometimes saved with a .xls
    extension). Has a few metadata preamble rows before the real header row,
    which starts with 'Keyword'. openpyxl validates by file extension, so if
    the file is named .xls but is actually xlsx-zip content, copy/read it as
    bytes via a .xlsx-named temp path first.
    """
    if not path.exists():
        return None
    read_path = path
    if path.suffix.lower() != ".xlsx":
        import shutil
        import tempfile
        tmp = Path(tempfile.gettempdir()) / (path.stem + "__aio_copy.xlsx")
        shutil.copyfile(path, tmp)
        read_path = tmp
    wb = load_workbook(read_path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and str(r[0]).strip() == "Keyword")
    header = [str(c) if c is not None else "" for c in rows[header_idx]]
    data = rows[header_idx + 1:]
    df = pd.DataFrame(data, columns=header)
    return df.dropna(how="all")


def build(source_dir: Path, out_path: Path, brand_domain: str, report_month: str):
    wb = Workbook()
    wb.remove(wb.active)

    # ---- README ----
    ws = wb.create_sheet("README")
    ws.sheet_properties.tabColor = "548235"
    for i, (text, font, _) in enumerate(README_LINES, start=1):
        cell = ws.cell(row=i, column=1, value=text)
        cell.font = font or BODY_FONT
    ws.column_dimensions["A"].width = 95

    # ---- Settings ----
    ws = wb.create_sheet("Settings")
    ws.sheet_properties.tabColor = "548235"
    ws.append(["Setting", "Value", "Notes"])
    style_header_row(ws, HEADER_FILL)
    rows = [
        ("Brand_Domain", brand_domain, "Your domain, lowercase, no protocol/www (used to flag Is_Brand / Brand_Mentioned)"),
        ("Report_Month", report_month, "Update this every month BEFORE running update_template.py, format YYYY-MM"),
        ("Last_Updated", "", "Set automatically by update_template.py"),
        ("Brand_Display_Name", "Atlas Copco", "Shown in the dashboard header/sidebar"),
    ]
    for r in rows:
        ws.append(r)
    autosize(ws, 3, 30)

    # ---- Mapping_Reference ----
    ws = wb.create_sheet("Mapping_Reference")
    ws.sheet_properties.tabColor = "548235"
    ws.append(["RAW Sheet", "Source / Export", "Feeds Clean Sheet(s)", "Transform Notes"])
    style_header_row(ws, HEADER_FILL)
    for r in MAPPING_REFERENCE:
        ws.append(r)
    widths = [22, 45, 28, 70]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # ---- RAW sheets ----
    aio = load_aio_xlsx(source_dir / "AI-overviews-SEMRUSH.xls")
    add_raw_sheet_from_df(wb, "RAW_AIO_Rankings", aio) if aio is not None else add_empty_raw_sheet(
        wb, "RAW_AIO_Rankings", ["Keyword"])

    brand = load_csv(source_dir / "brand_topics_atlascopco_com_us.csv")
    add_raw_sheet_from_df(wb, "RAW_Brand_Topics", brand) if brand is not None else add_empty_raw_sheet(
        wb, "RAW_Brand_Topics", ["name", "country", "visibility", "difficulty", "mentions", "volume", "volume_trend", "intents"])

    # No topic-level Gap Topics export for this client -- always empty.
    add_empty_raw_sheet(
        wb, "RAW_Gap_Topics", ["name", "country", "visibility", "difficulty", "mentions", "gap_mentions", "volume", "volume_trend", "intents"])

    prompts = load_csv(source_dir / "prompts_atlascopco_com_us.csv")
    add_raw_sheet_from_df(wb, "RAW_Prompts", prompts) if prompts is not None else add_empty_raw_sheet(
        wb, "RAW_Prompts", ["prompt", "country", "llm", "categories", "topic_id", "topic_name", "topic_volume", "topic_intents", "topic_visibility", "topic_difficulty", "brief_response", "mentioned_brands_count", "sources_count"])

    # No Sources export for this client -- always empty (dashboard hides the tab).
    add_empty_raw_sheet(
        wb, "RAW_Sources", ["url", "country", "prompts_count"])

    chatgpt = load_csv(source_dir / "chatgpt_results_uae.csv")
    add_raw_sheet_from_df(wb, "RAW_ChatGPT_Results", chatgpt) if chatgpt is not None else add_empty_raw_sheet(
        wb, "RAW_ChatGPT_Results", ["keyword", "answer", "source_urls"])

    ga4_path = source_dir / "AI-Traffic-analytics.xlsx"
    ga4 = pd.read_excel(ga4_path, sheet_name="Sheet1") if ga4_path.exists() else None
    add_raw_sheet_from_df(wb, "RAW_GA4_Traffic", ga4) if ga4 is not None else add_empty_raw_sheet(
        wb, "RAW_GA4_Traffic", ["Month", "ChatGPT", "Gemini", "Copilot", "Total AI (Top 3)"])

    gsc_path = source_dir / "gsc-AIO-Comparison.xlsx"
    if not gsc_path.exists():
        gsc_path = source_dir / "AI-Performanc-GSC.xlsx"
    if gsc_path.exists():
        gsc_pages = pd.read_excel(gsc_path, sheet_name="Pages")
        gsc_countries = pd.read_excel(gsc_path, sheet_name="Countries")
        gsc_devices = pd.read_excel(gsc_path, sheet_name="Devices")
    else:
        gsc_pages = gsc_countries = gsc_devices = None
    add_raw_sheet_from_df(wb, "RAW_GSC_Pages", gsc_pages) if gsc_pages is not None else add_empty_raw_sheet(
        wb, "RAW_GSC_Pages", ["Top pages", "Impressions"])
    add_raw_sheet_from_df(wb, "RAW_GSC_Countries", gsc_countries) if gsc_countries is not None else add_empty_raw_sheet(
        wb, "RAW_GSC_Countries", ["Country", "Impressions"])
    add_raw_sheet_from_df(wb, "RAW_GSC_Devices", gsc_devices) if gsc_devices is not None else add_empty_raw_sheet(
        wb, "RAW_GSC_Devices", ["Device", "Impressions"])

    # ---- Clean sheets (headers only; populated by update_template.py) ----
    for name in CLEAN_SCHEMAS:
        add_clean_sheet(wb, name)

    wb.save(out_path)
    print(f"Created {out_path}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--source-dir", default=".")
    ap.add_argument("--out", default="AI_Visibility_Tracking_Template.xlsx")
    ap.add_argument("--brand-domain", default="atlascopco.com")
    ap.add_argument("--report-month", default=None)
    args = ap.parse_args()
    import datetime as dt
    report_month = args.report_month or dt.date.today().strftime("%Y-%m")
    build(Path(args.source_dir), Path(args.out), args.brand_domain.lower(), report_month)
