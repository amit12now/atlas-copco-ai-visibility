#!/usr/bin/env python3
"""
update_template.py  (Atlas Copco build)
-----------------------------------------
Monthly refresh script for AI_Visibility_Tracking_Template.xlsx.

WORKFLOW (run this every month):
  1. Export fresh data from Semrush (AI Overview tracking, Brand AI-visibility
     topics, Prompts), ChatGPT checker, GA4, and Google Search Console.
  2. Open AI_Visibility_Tracking_Template.xlsx. For each RAW_* tab, clear the old
     rows below the header and paste the new export in (matching the existing
     header row).
  3. On the Settings tab, set Report_Month to the current period, e.g. "2026-07".
  4. Save the file.
  5. Run:  python update_template.py "AI_Visibility_Tracking_Template.xlsx"
     (or just `python update_template.py` if run from the same folder as the file)

This script reads every RAW_* sheet, cleans/normalizes it, and upserts the result
into the matching Clean_* sheet (replacing any existing rows for the same
Report_Month so it is safe to re-run). The Streamlit dashboard reads ONLY the
Clean_* sheets, never the RAW_* sheets.

NOTES ON THIS CLIENT'S EXPORT SHAPES (vs. the original template):
  - RAW_AIO_Rankings is a DAILY position-tracking export (one column block per
    domain per calendar day across the report period) rather than a single
    monthly snapshot. transform_aio() unpivots every domain/day block and then
    averages Position/Visibility across the days into one row per Keyword x
    Domain for Report_Month -- Position_Change/Visibility_Change become the
    period-long Semrush "difference" columns (one per domain) rather than a
    prior-month delta.
  - RAW_GA4_Traffic has real datetime Month values for past months and a
    string like "June 2026*" for the current (partial) month, no "(domain)"
    suffixes on the platform columns, and an extra "Total AI (Top 3)" column
    that must be dropped before unpivoting.
  - RAW_GSC_Pages/Countries/Devices (gsc-AIO-Comparison.xlsx) come from Search
    Console's AI Overviews / AI Mode reporting (NOT classic web search --
    "AIO" = AI Overview), filtered to the /compressors site section. They
    have two columns shaped like "6/1/26 - 6/26/26 Impressions" /
    "5/1/26 - 5/26/26 Impressions" -- a real current-period-vs-prior-period
    comparison, just not labeled "Last 28" / "Previous 28" like the original
    template's source. transform_gsc() parses the start date in each such
    column to tell current from prior. Falls back to a single "Impressions"
    column (Change_Pct left blank) if an export without a comparison is used
    instead.
"""
import sys
import re
import datetime as dt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DEFAULT_PATH = "AI_Visibility_Tracking_Template.xlsx"

HEADER_FILL = PatternFill("solid", start_color="1F3864", end_color="1F3864")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)

INTENT_MAP = {
    "task": "Intent_Task",
    "informational": "Intent_Informational",
    "navigational": "Intent_Navigational",
    "commercial": "Intent_Commercial",
    "transactional": "Intent_Transactional",
}
INTENT_COLS = list(INTENT_MAP.values())


# --------------------------------------------------------------------------
# Helpers
# --------------------------------------------------------------------------

def read_sheet(path, sheet_name):
    try:
        df = pd.read_excel(path, sheet_name=sheet_name)
    except ValueError:
        return pd.DataFrame()
    return df.dropna(how="all")


def get_settings(path):
    df = read_sheet(path, "Settings")
    settings = {}
    if not df.empty and {"Setting", "Value"}.issubset(df.columns):
        for _, row in df.iterrows():
            settings[str(row["Setting"]).strip()] = row["Value"]
    brand_domain = str(settings.get("Brand_Domain", "")).strip().lower()
    report_month = str(settings.get("Report_Month", "")).strip()
    if not re.match(r"^\d{4}-\d{2}$", report_month):
        raise ValueError(
            f"Settings!Report_Month must look like 'YYYY-MM' (got '{report_month}')"
        )
    return brand_domain, report_month


def split_intents(series):
    """Parse 'task:1;informational:46;...' strings into 5 numeric columns."""
    out = pd.DataFrame(0, index=series.index, columns=INTENT_COLS)
    for idx, val in series.items():
        if pd.isna(val):
            continue
        for chunk in str(val).split(";"):
            if ":" not in chunk:
                continue
            key, _, num = chunk.partition(":")
            col = INTENT_MAP.get(key.strip().lower())
            if col:
                out.at[idx, col] = pd.to_numeric(num, errors="coerce")
    return out


def to_num(series):
    return pd.to_numeric(series, errors="coerce")


def upsert(path, sheet_name, new_df, key_cols):
    """Replace rows in `sheet_name` whose key_cols match new_df's key_cols, then append new_df."""
    existing = read_sheet(path, sheet_name)
    if not existing.empty and all(k in existing.columns for k in key_cols):
        key_vals = new_df[key_cols].drop_duplicates()
        merge_key = key_vals.apply(lambda r: tuple(r), axis=1).tolist()
        existing_keys = existing[key_cols].apply(lambda r: tuple(r), axis=1)
        existing = existing[~existing_keys.isin(merge_key)]
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df
    return combined


# --------------------------------------------------------------------------
# Transform: AI Overview Rankings (Semrush position tracking, AIO filter)
#   Atlas Copco's export is a DAILY snapshot per domain across the report
#   period (one position/type/landing block per domain per calendar day),
#   plus ONE period-long "difference" column per domain (no per-day
#   visibility / visibility_difference siblings). We unpivot every day/domain
#   block, then collapse to one row per Keyword x Domain for Report_Month by
#   averaging Position across the days it actually ranked.
# --------------------------------------------------------------------------

def transform_aio(raw, brand_domain, report_month):
    if raw.empty:
        return pd.DataFrame()
    pos_pattern = re.compile(r"^\*\.(?P<domain>[^/]+)/\*_(?P<date>\d{8})$")
    blocks = []
    for col in raw.columns:
        m = pos_pattern.match(str(col))
        if m:
            blocks.append((m.group("domain"), m.group("date"), col))

    rows = []
    for domain, date, pos_col in blocks:
        type_col = f"{pos_col}_type"
        landing_col = f"{pos_col}_landing"
        diff_col = f"*.{domain}/*_difference"

        block = pd.DataFrame({
            "Keyword": raw.get("Keyword"),
            "Domain": domain,
            "Date": date,
            "Position": raw.get(pos_col).replace("-", np.nan) if pos_col in raw else np.nan,
            "Result_Type": raw.get(type_col) if type_col in raw else np.nan,
            "Landing_URL": raw.get(landing_col) if landing_col in raw else np.nan,
            "Position_Change": raw.get(diff_col).replace("-", np.nan) if diff_col in raw else np.nan,
            "Tags": raw.get("Tags"),
            "Intent_Code": raw.get("Intents"),
            "CPC": raw.get("CPC"),
            "Search_Volume": raw.get("Search Volume"),
            "Keyword_Difficulty": raw.get("Keyword Difficulty"),
        })
        rows.append(block)

    long_df = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()
    if long_df.empty:
        return long_df
    long_df["Position"] = to_num(long_df["Position"])
    long_df["Position_Change"] = to_num(long_df["Position_Change"])
    long_df["CPC"] = to_num(long_df["CPC"])
    long_df["Search_Volume"] = to_num(long_df["Search_Volume"])
    long_df["Keyword_Difficulty"] = to_num(long_df["Keyword_Difficulty"])

    # Collapse the daily snapshots into one row per Keyword x Domain: average
    # Position across days actually ranked (NaN days excluded automatically),
    # keep the most recent non-null Result_Type/Landing_URL, and keep the
    # period-long Position_Change as-is (it's already one value per domain).
    long_df = long_df.sort_values("Date")
    agg = long_df.groupby(["Keyword", "Domain"]).agg(
        Position=("Position", "mean"),
        Result_Type=("Result_Type", "last"),
        Landing_URL=("Landing_URL", "last"),
        Position_Change=("Position_Change", "last"),
        Tags=("Tags", "last"),
        Intent_Code=("Intent_Code", "last"),
        CPC=("CPC", "last"),
        Search_Volume=("Search_Volume", "last"),
        Keyword_Difficulty=("Keyword_Difficulty", "last"),
    ).reset_index()

    agg["Visibility_Pct"] = np.nan
    agg["Visibility_Change"] = np.nan
    agg["Is_Brand"] = agg["Domain"].str.lower() == brand_domain
    agg["Report_Month"] = report_month
    cols = ["Report_Month", "Keyword", "Domain", "Is_Brand", "Position", "Visibility_Pct",
            "Result_Type", "Landing_URL", "Position_Change", "Visibility_Change",
            "Search_Volume", "Keyword_Difficulty", "CPC", "Intent_Code", "Tags"]
    return agg[cols]


# --------------------------------------------------------------------------
# Transform: Brand Topics (Semrush AI visibility)
# --------------------------------------------------------------------------

def transform_brand_topics(raw, report_month):
    if raw.empty:
        return pd.DataFrame()
    out = pd.DataFrame({
        "Report_Month": report_month,
        "Topic": raw.get("name"),
        "Country": raw.get("country"),
        "Visibility": to_num(raw.get("visibility")),
        "Difficulty": to_num(raw.get("difficulty")),
        "Mentions": to_num(raw.get("mentions")),
        "Search_Volume": to_num(raw.get("volume")),
        "Volume_Trend": raw.get("volume_trend"),
    })
    out = pd.concat([out, split_intents(raw.get("intents", pd.Series(dtype=str)))], axis=1)
    return out


def transform_gap_topics(raw, report_month):
    """No topic-level Gap Topics export for this client -- always returns
    empty frames. Kept so the script structure matches the reusable template."""
    if raw.empty:
        return pd.DataFrame(), pd.DataFrame()
    out = pd.DataFrame({
        "Report_Month": report_month,
        "Topic": raw.get("name"),
        "Country": raw.get("country"),
        "Visibility": to_num(raw.get("visibility")),
        "Difficulty": to_num(raw.get("difficulty")),
        "Brand_Mentions": to_num(raw.get("mentions")),
        "Search_Volume": to_num(raw.get("volume")),
        "Volume_Trend": raw.get("volume_trend"),
    })
    out = pd.concat([out, split_intents(raw.get("intents", pd.Series(dtype=str)))], axis=1)

    detail_rows = []
    for _, row in raw.iterrows():
        packed = row.get("gap_mentions")
        if pd.isna(packed):
            continue
        for chunk in str(packed).split(";"):
            if ":" not in chunk:
                continue
            domain, _, mentions = chunk.partition(":")
            detail_rows.append({
                "Report_Month": report_month,
                "Topic": row.get("name"),
                "Competitor_Domain": domain.strip(),
                "Competitor_Mentions": pd.to_numeric(mentions, errors="coerce"),
            })
    detail = pd.DataFrame(detail_rows)
    return out, detail


# --------------------------------------------------------------------------
# Transform: Prompts, Sources, ChatGPT results
# --------------------------------------------------------------------------

def transform_prompts(raw, report_month):
    if raw.empty:
        return pd.DataFrame()
    out = pd.DataFrame({
        "Report_Month": report_month,
        "Prompt": raw.get("prompt"),
        "Country": raw.get("country"),
        "LLM": raw.get("llm"),
        "Category": raw.get("categories"),
        "Topic_ID": raw.get("topic_id").astype(str) if "topic_id" in raw else None,
        "Topic_Name": raw.get("topic_name"),
        "Topic_Volume": to_num(raw.get("topic_volume")),
        "Topic_Visibility": to_num(raw.get("topic_visibility")),
        "Topic_Difficulty": to_num(raw.get("topic_difficulty")),
        "Mentioned_Brands_Count": to_num(raw.get("mentioned_brands_count")),
        "Sources_Count": to_num(raw.get("sources_count")),
        "Brief_Response": raw.get("brief_response"),
    })
    intents = split_intents(raw.get("topic_intents", pd.Series(dtype=str)))
    intents.columns = [f"Topic_{c}" for c in intents.columns]
    return pd.concat([out, intents], axis=1)


def transform_sources(raw, report_month):
    """No Sources export for this client -- always returns empty. Kept so the
    script structure matches the reusable template."""
    if raw.empty:
        return pd.DataFrame()
    return pd.DataFrame({
        "Report_Month": report_month,
        "URL": raw.get("url"),
        "Country": raw.get("country"),
        "Prompts_Count": to_num(raw.get("prompts_count")),
    })


def transform_chatgpt(raw, brand_domain, report_month):
    if raw.empty:
        return pd.DataFrame()
    source_urls = raw.get("source_urls").fillna("")
    url_count = source_urls.apply(lambda s: len([u for u in s.split("|") if u.strip()]))
    brand_mentioned = (
        raw.get("answer").fillna("").str.contains(brand_domain, case=False)
        | source_urls.str.contains(brand_domain, case=False)
    )
    return pd.DataFrame({
        "Report_Month": report_month,
        "Keyword": raw.get("keyword"),
        "Brand_Mentioned": brand_mentioned,
        "Source_URL_Count": url_count,
        "Answer": raw.get("answer"),
        "Source_URLs": raw.get("source_urls"),
    })


# --------------------------------------------------------------------------
# Transform: GA4 AI referral traffic (wide -> long)
#   Month is a real datetime for closed months, and a string like
#   "June 2026*" for the current in-progress month (the trailing "*" is the
#   to-date marker). Platform columns have no "(domain)" suffix, and there is
#   a redundant "Total AI (Top 3)" column to drop.
# --------------------------------------------------------------------------

MONTH_NAMES = ["January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def transform_ga4(raw, report_month):
    if raw.empty:
        return pd.DataFrame()
    raw = raw[raw["Month"].astype(str).str.strip().str.lower() != "total"].copy()
    raw = raw.drop(columns=[c for c in raw.columns if str(c).strip().lower().startswith("total")])

    def parse_month(label):
        if isinstance(label, (dt.datetime, pd.Timestamp)):
            return label.strftime("%B"), label.year, False
        s = str(label).strip()
        is_todate = s.endswith("*")
        s = s.rstrip("*").strip()
        m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", s)
        if m:
            return m.group(1), int(m.group(2)), is_todate
        return s, None, is_todate

    parsed = raw["Month"].apply(parse_month)
    raw["Month_Label"] = [p[0] for p in parsed]
    raw["Year"] = [p[1] for p in parsed]
    raw["Is_To_Date"] = [p[2] for p in parsed]
    month_nums = [MONTH_NAMES.index(m) + 1 if m in MONTH_NAMES else None for m in raw["Month_Label"]]
    raw["Year_Month"] = [
        f"{y}-{m:02d}" if y and m else None for y, m in zip(raw["Year"], month_nums)
    ]

    platform_cols = [c for c in raw.columns if c not in
                      ("Month", "Month_Label", "Year", "Is_To_Date", "Year_Month")]
    long_rows = []
    for col in platform_cols:
        m = re.match(r"^(?P<name>.+?)\s*\((?P<domain>.+?)\)\s*$", str(col))
        platform = m.group("name").strip() if m else str(col)
        platform_domain = m.group("domain").strip() if m else ""
        sub = raw[["Year_Month", "Month_Label", "Is_To_Date"]].copy()
        sub["Platform"] = platform
        sub["Platform_Domain"] = platform_domain
        sub["Sessions"] = to_num(raw[col])
        long_rows.append(sub)
    out = pd.concat(long_rows, ignore_index=True)
    return out[["Year_Month", "Month_Label", "Is_To_Date", "Platform", "Platform_Domain", "Sessions"]]


# --------------------------------------------------------------------------
# Transform: GSC AI Overviews / AI Mode (Pages / Countries / Devices -> one long table)
#   gsc-AIO-Comparison.xlsx is exported from Search Console's AI Overviews /
#   AI Mode reporting ("AIO" = AI Overview), not the classic web-search
#   Performance report. It has two columns shaped like
#   "6/1/26 - 6/26/26 Impressions" (current period) and
#   "5/1/26 - 5/26/26 Impressions" (prior period) -- a real comparison, just
#   not labeled "Last 28" / "Previous 28" like the original template's
#   source. We detect those date-range columns and parse their start dates to
#   tell current from prior. If only a single "Impressions" column exists
#   (an older/simpler export), Impressions_Prev_28d / Change_Pct are left
#   blank -- the dashboard already renders gracefully without them (falls
#   back to a flat "■" trend indicator).
# --------------------------------------------------------------------------

DATE_RANGE_COL_RE = re.compile(
    r"^(?P<start>\d{1,2}/\d{1,2}/\d{2,4})\s*-\s*(?P<end>\d{1,2}/\d{1,2}/\d{2,4})\s+Impressions$"
)


def _parse_date_range_cols(df):
    """Find columns shaped like 'M/D/YY - M/D/YY Impressions' (a period vs.
    period comparison export) and return them sorted most-recent-first by
    parsing each column's start date."""
    found = []
    for col in df.columns:
        m = DATE_RANGE_COL_RE.match(str(col).strip())
        if not m:
            continue
        start_str = m.group("start")
        start = None
        for fmt in ("%m/%d/%y", "%m/%d/%Y"):
            try:
                start = dt.datetime.strptime(start_str, fmt)
                break
            except ValueError:
                continue
        if start is not None:
            found.append((start, col))
    found.sort(key=lambda t: t[0], reverse=True)
    return found


def transform_gsc(pages, countries, devices, report_month):
    frames = []
    specs = [(pages, "Page", "Top pages"),
             (countries, "Country", "Country"),
             (devices, "Device", "Device")]
    for df, dim_type, label_col in specs:
        if df.empty:
            continue
        last_candidates = [c for c in df.columns if "Last 28" in c]
        prev_candidates = [c for c in df.columns if "Previous 28" in c]
        date_range_cols = _parse_date_range_cols(df)
        if last_candidates:
            last_vals = to_num(df.get(last_candidates[0]))
            prev_vals = (to_num(df.get(prev_candidates[0])) if prev_candidates
                         else pd.Series([np.nan] * len(df), index=df.index))
        elif len(date_range_cols) >= 2:
            last_vals = to_num(df.get(date_range_cols[0][1]))
            prev_vals = to_num(df.get(date_range_cols[1][1]))
        elif len(date_range_cols) == 1:
            last_vals = to_num(df.get(date_range_cols[0][1]))
            prev_vals = pd.Series([np.nan] * len(df), index=df.index)
        else:
            last_vals = to_num(df.get("Impressions"))
            prev_vals = pd.Series([np.nan] * len(df), index=df.index)
        sub = pd.DataFrame({
            "Report_Month": report_month,
            "Dimension_Type": dim_type,
            "Dimension_Value": df.get(label_col),
            "Impressions_Last_28d": last_vals,
            "Impressions_Prev_28d": prev_vals,
        })
        sub["Change_Pct"] = np.where(
            sub["Impressions_Prev_28d"] > 0,
            (sub["Impressions_Last_28d"] - sub["Impressions_Prev_28d"]) / sub["Impressions_Prev_28d"] * 100,
            np.nan,
        )
        frames.append(sub)
    return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()


# --------------------------------------------------------------------------
# Write-back with light formatting
# --------------------------------------------------------------------------

def write_clean_sheet(path, sheet_name, df):
    if df.empty:
        return
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(path)
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT
    for i, col in enumerate(df.columns, start=1):
        width = max(12, min(45, int(df[col].astype(str).str.len().clip(upper=60).mean()) + 6))
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    wb.save(path)


def set_last_updated(path):
    wb = load_workbook(path)
    ws = wb["Settings"]
    for row in ws.iter_rows():
        if row[0].value == "Last_Updated":
            row[1].value = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
            break
    wb.save(path)


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main(path):
    brand_domain, report_month = get_settings(path)
    print(f"Brand domain: {brand_domain} | Report month: {report_month}")

    raw_aio = read_sheet(path, "RAW_AIO_Rankings")
    raw_brand = read_sheet(path, "RAW_Brand_Topics")
    raw_gap = read_sheet(path, "RAW_Gap_Topics")
    raw_prompts = read_sheet(path, "RAW_Prompts")
    raw_sources = read_sheet(path, "RAW_Sources")
    raw_chatgpt = read_sheet(path, "RAW_ChatGPT_Results")
    raw_ga4 = read_sheet(path, "RAW_GA4_Traffic")
    raw_gsc_pages = read_sheet(path, "RAW_GSC_Pages")
    raw_gsc_countries = read_sheet(path, "RAW_GSC_Countries")
    raw_gsc_devices = read_sheet(path, "RAW_GSC_Devices")

    aio = transform_aio(raw_aio, brand_domain, report_month)
    brand_topics = transform_brand_topics(raw_brand, report_month)
    gap_topics, gap_detail = transform_gap_topics(raw_gap, report_month)
    prompts = transform_prompts(raw_prompts, report_month)
    sources = transform_sources(raw_sources, report_month)
    chatgpt = transform_chatgpt(raw_chatgpt, brand_domain, report_month)
    ga4 = transform_ga4(raw_ga4, report_month)
    gsc = transform_gsc(raw_gsc_pages, raw_gsc_countries, raw_gsc_devices, report_month)

    jobs = [
        ("Clean_AIO_Rankings", aio, ["Report_Month"]),
        ("Clean_Brand_Topics", brand_topics, ["Report_Month"]),
        ("Clean_Gap_Topics", gap_topics, ["Report_Month"]),
        ("Clean_Gap_Topics_Competitors", gap_detail, ["Report_Month"]),
        ("Clean_Prompts", prompts, ["Report_Month"]),
        ("Clean_Sources", sources, ["Report_Month"]),
        ("Clean_ChatGPT_Results", chatgpt, ["Report_Month"]),
        ("Clean_GA4_Traffic", ga4, ["Year_Month", "Platform"]),
        ("Clean_GSC_Performance", gsc, ["Report_Month", "Dimension_Type"]),
    ]
    for sheet_name, new_df, key_cols in jobs:
        if new_df.empty:
            print(f"  skip {sheet_name}: no RAW data found")
            continue
        combined = upsert(path, sheet_name, new_df, key_cols)
        write_clean_sheet(path, sheet_name, combined)
        print(f"  {sheet_name}: {len(new_df)} rows this run, {len(combined)} total rows")

    set_last_updated(path)
    print("Done.")


if __name__ == "__main__":
    target = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_PATH
    main(target)
